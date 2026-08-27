import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import F, Q
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

from apps.users.decorators import (
    admin_or_manager_or_staff_required,
    admin_or_manager_required,
    admin_required,
)

from .forms import ClientForm, ClientPhotoForm, ImportClientsForm, SevenHillsRegistrationForm
from .models import Client, ClientProfilePicture, SevenHillsRegistration


# =================================== Fetch and display all clients details ===================================
@login_required
@admin_or_manager_or_staff_required
def client_list(request):
    base_queryset = Client.objects.prefetch_related(
        "loans__documents", "profile_pictures"
    ).order_by(F("reg_number").asc(nulls_last=True), "id")
    queryset = base_queryset

    search_query = request.GET.get("search", "").strip()
    if search_query:
        queryset = queryset.filter(
            Q(full_name__icontains=search_query)
            | Q(reg_number__icontains=search_query)
            | Q(mobile_telephone__icontains=search_query)
            | Q(email__icontains=search_query)
        )

    paginator = Paginator(queryset, 20)  # Show 20 records per page
    page = request.GET.get("page")

    try:
        records = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        records = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        records = paginator.page(paginator.num_pages)

    return render(
        request,
        "client/client_list.html",
        {
            "records": records,
            "table_title": "Clients List",
            "search_query": search_query,
            "total_clients": base_queryset.count(),
            "clients_with_phone": base_queryset.exclude(mobile_telephone__isnull=True)
            .exclude(mobile_telephone="")
            .count(),
            "clients_with_email": base_queryset.exclude(email__isnull=True)
            .exclude(email="")
            .exclude(email="no-email@example.com")
            .count(),
        },
    )


# =================================== Upload Client Photo ===================================


@login_required
@transaction.atomic
@admin_or_manager_or_staff_required
def upload_client_photo(request):
    clients = Client.objects.order_by("full_name", "id")
    form = ClientPhotoForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        client = get_object_or_404(clients, pk=request.POST.get("id"))
        remove_requested = request.POST.get("remove_picture") == "1"
        uploaded_picture = form.cleaned_data.get("picture")

        if remove_requested and not uploaded_picture:
            ClientProfilePicture.objects.filter(client=client, is_current=True).update(
                is_current=False
            )
            client.picture = None
            client.save(update_fields=["picture", "updated_at"])
            messages.success(
                request,
                f"Current profile picture removed from {client.full_name}.",
                extra_tags="bg-success",
            )
            return redirect("upload_client_photo")

        if uploaded_picture:
            ClientProfilePicture.objects.filter(client=client, is_current=True).update(
                is_current=False
            )
            photo = ClientProfilePicture.objects.create(
                client=client,
                picture=uploaded_picture,
                is_current=True,
            )
            client.picture = photo.picture
            client.save(update_fields=["picture", "updated_at"])
            messages.success(
                request,
                f"Photo updated for {client.full_name}.",
                extra_tags="bg-success",
            )
            return redirect("upload_client_photo")

        form.add_error("picture", "Take a photo, choose one from the gallery, or remove the current picture.")

    if request.method == "POST":
        messages.error(
            request,
            "The photo could not be uploaded. Check the selected client and photo.",
            extra_tags="bg-danger",
        )

    return render(
        request,
        "client/client_photo.html",
        {
            "form": form,
            "clients": clients,
            "form_name": "Upload Client Photo",
            "allow_current_photo_removal": True,
        },
    )


@login_required
@admin_or_manager_required
@transaction.atomic
def delete_client_profile_picture(request, pk):
    if request.method != "POST":
        messages.error(request, "Use the Delete button to remove a client photo.")
        return redirect("client_list")

    photo = get_object_or_404(ClientProfilePicture.objects.select_related("client"), pk=pk)
    client = photo.client
    was_current = photo.is_current or str(client.picture) == str(photo.picture)
    photo.delete()

    if was_current:
        replacement = ClientProfilePicture.objects.filter(client=client).first()
        if replacement:
            replacement.is_current = True
            replacement.save(update_fields=["is_current"])
            client.picture = replacement.picture
        else:
            client.picture = None
        client.save(update_fields=["picture", "updated_at"])

    messages.info(request, f"Photo removed from {client.full_name}.", extra_tags="bg-danger")
    return redirect("client_list")


# =================================== Register Client  ===================================


@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def register_client(request):
    if request.method == "POST":
        form = ClientForm(request.POST, request.FILES)

        if form.is_valid():
            client = form.save()
            if client.picture:
                ClientProfilePicture.objects.filter(client=client, is_current=True).update(
                    is_current=False
                )
                ClientProfilePicture.objects.create(
                    client=client,
                    picture=client.picture,
                    is_current=True,
                )
            messages.success(
                request, "Record saved successfully!", extra_tags="bg-success"
            )
            return redirect("register_client")
        else:
            # Display an error message if the form is not valid
            messages.error(
                request,
                "There was an error saving the record. Please check the form for errors.",
                extra_tags="bg-danger",
            )

    else:
        form = ClientForm()

    return render(
        request,
        "client/client_register.html",
        {"form_name": "Client Registration", "form": form},
    )


# =================================== Update client data ===================================
@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def update_client(request, pk, template_name="client/client_update.html"):
    try:
        client_record = Client.objects.get(pk=pk)
    except Client.DoesNotExist:
        messages.error(request, "Client record not found!", extra_tags="bg-danger")
        return redirect("client_list")  # Or a relevant error page

    if request.method == "POST":
        form = ClientForm(request.POST, request.FILES, instance=client_record)
        if form.is_valid():
            uploaded_picture = request.FILES.get("picture")
            remove_picture = (
                request.POST.get("remove_picture") == "1" and not uploaded_picture
            )
            client = form.save(commit=False)
            if remove_picture:
                client.picture = None
            client.save()

            if uploaded_picture and client.picture:
                ClientProfilePicture.objects.filter(client=client, is_current=True).update(
                    is_current=False
                )
                ClientProfilePicture.objects.create(
                    client=client,
                    picture=client.picture,
                    is_current=True,
                )
            elif remove_picture:
                ClientProfilePicture.objects.filter(client=client, is_current=True).update(
                    is_current=False
                )

            messages.success(
                request, "Client record updated successfully!", extra_tags="bg-success"
            )
            return redirect("client_list")
    else:
        # Pre-populate the form with existing data
        form = ClientForm(instance=client_record)

    context = {
        "form_name": "Client Registration",
        "form": form,
        "current_photo_url": client_record.picture.url if client_record.picture else "",
        "photo_subject": client_record.full_name,
    }
    return render(request, template_name, context)


# =================================== Delete selected client ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def delete_client(request, pk):
    records = Client.objects.get(id=pk)
    records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("client_list"))


# =================================== Process and Import Excel data ===================================
@login_required
@admin_required
@transaction.atomic
def import_client_data(request):
    if request.method == "POST":
        form = ImportClientsForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES.get("excel_file")
            if excel_file and excel_file.name.endswith(".xlsx"):
                try:
                    # Call process_and_import_data function
                    errors = process_and_import_data(excel_file)
                    if errors:
                        for error in errors:
                            messages.error(request, error, extra_tags="bg-danger")
                    else:
                        messages.success(
                            request,
                            "Data imported successfully!",
                            extra_tags="bg-success",
                        )
                except Exception as e:
                    messages.error(
                        request, f"Error importing data: {e}", extra_tags="bg-danger"
                    )
                return redirect("client_list")
            else:
                messages.error(
                    request, "Please upload a valid Excel file.", extra_tags="bg-danger"
                )
    else:
        form = ImportClientsForm()
    return render(
        request,
        "client/bulk_import.html",
        {"form_name": "Import Clients - Excel", "form": form},
    )


# Function to import Excel data
def process_and_import_data(excel_file):
    errors = []
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            fname = row[0].value
            picture = row[1].value
            reg_number = row[2].value
            mobile_telephone = row[3].value
            if fname is not None:
                try:
                    Client.objects.create(
                        full_name=fname,
                        picture=picture,
                        reg_number=reg_number,
                        mobile_telephone=mobile_telephone,
                    )
                except Exception as e:
                    errors.append(f"Error on row {row_num}: {e}")
            else:
                errors.append(f"Missing full name on row {row_num}")
    except Exception as e:
        errors.append(f"Failed to process the Excel file: {e}")
    return errors


# =================================== Delete all records at once ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def delete_confirm(request):
    if request.method == "POST":
        Client.objects.all().delete()
        messages.info(request, "All records deleted!", extra_tags="bg-danger")
        return HttpResponseRedirect(reverse("client_list"))


# =================================== seven_hills registration ===================================
@login_required
@transaction.atomic
def seven_hills_registration_view(request):
    if request.method == "POST":
        form = SevenHillsRegistrationForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            messages.success(
                request, "Record saved successfully!", extra_tags="bg-success"
            )
            return redirect("seven_hills_registration")
        else:
            # Display error messages if the form is invalid
            messages.error(
                request,
                "There was an error saving the record. Please check the form for errors.",
                extra_tags="bg-danger",
            )
    else:
        form = SevenHillsRegistrationForm()

    context = {
        "form_name": "Seven Hills Registration Form",
        "form": form,
    }

    return render(request, "client/seven_hills_register.html", context)


# =================================== Fetch and display all Seven Hills Registration details ===================================
@login_required
def seven_hills_list(request):
    # Fetch all records
    queryset = SevenHillsRegistration.objects.all().order_by("id")

    # Apply search filter
    search_query = request.GET.get("search")
    if search_query:
        queryset = queryset.filter(
            Q(full_name__icontains=search_query)
            | Q(residence__icontains=search_query)
            | Q(services_interested__icontains=search_query)
            | Q(ministry_groups__icontains=search_query)
        )
        if not queryset.exists():
            messages.info(request, "No results found for your search.")

    # Paginate the filtered queryset
    paginator = Paginator(queryset, 100)
    page = request.GET.get("page")

    try:
        records = paginator.page(page)
    except PageNotAnInteger:
        records = paginator.page(1)
    except EmptyPage:
        records = paginator.page(paginator.num_pages)

    # Pass both the full queryset and paginated records to the template
    return render(
        request,
        "client/seven_hills_list.html",
        {
            "records": records,  # Paginated records for display
            "table_title": "Seven Hills Members List",
            "queryset": queryset,  # Full queryset if needed elsewhere
        },
    )


# =================================== Update Seven Hills data ===================================
@login_required
@transaction.atomic
def update_seven_hills(request, pk, template_name="client/seven_hills_update.html"):
    try:
        record = SevenHillsRegistration.objects.get(pk=pk)
    except SevenHillsRegistration.DoesNotExist:
        messages.error(request, "Record not found!", extra_tags="bg-danger")
        return redirect("seven_hills_list")  # Or a relevant error page

    if request.method == "POST":
        form = SevenHillsRegistrationForm(request.POST, request.FILES, instance=record)
        if form.is_valid():
            form.save()

            messages.success(
                request, "Record updated successfully!", extra_tags="bg-success"
            )
            return redirect("seven_hills_list")
    else:
        # Pre-populate the form with existing data
        form = SevenHillsRegistrationForm(instance=record)

    context = {"form_name": "Seven Hills Update", "form": form}
    return render(request, template_name, context)


# =================================== Delete selected Seven Hills ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def delete_seven_hills(request, pk):
    records = SevenHillsRegistration.objects.get(id=pk)
    records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("seven_hills_list"))


# =================================== Fetch and display selected member details ===================================
@login_required
def seven_hills_details(request, pk):
    record = SevenHillsRegistration.objects.get(pk=pk)
    age = record.calculate_age()

    context = {"table_title": "Profile Report", "record": record, "age": age}
    return render(request, "client/seven_hills_profile_rpt.html", context)
